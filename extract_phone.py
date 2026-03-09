import os
import re
import html
import io
from email import policy
from email.parser import BytesParser
from email.utils import parseaddr
from openpyxl import load_workbook, Workbook

# ✅ GUI
import tkinter as tk
from tkinter import filedialog, messagebox

try:
    import extract_msg
except ImportError:
    extract_msg = None

# ✅ OCR (optional)
OCR_AVAILABLE = True
try:
    from PIL import Image
    import pytesseract
except Exception:
    OCR_AVAILABLE = False


# =========================
# OCR CONFIG (EDIT IF NEEDED)
# =========================
TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"


# =========================
# Theme Colors
# =========================
SE_BLUE = "#0B4EA2"
SE_YELLOW = "#F7B500"
SE_BLACK = "#0B0B0B"
SE_WHITE = "#FFFFFF"
SE_DARK_CARD = "#141414"
SE_BORDER = "#2A2A2A"
SE_MUTED = "#AAAAAA"


# =========================
# Patterns / Filters
# =========================

EMAIL_REGEX = re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)

ON_WROTE_RE = re.compile(
    r"(\n\s*On\s.+?\bwrote:\s*\n)|"
    r"(\n\s*Le\s.+?\ba écrit\s*:?\s*\n)|"
    r"(\n\s*Am\s.+?\bschrieb\s*:?\s*\n)",
    re.IGNORECASE | re.DOTALL
)

THREAD_MARKERS = [
    r"\n\s*From:\s", r"\n\s*Sent:\s", r"\n\s*To:\s", r"\n\s*Subject:\s",
    r"\n\s*Von:\s", r"\n\s*Gesendet:\s", r"\n\s*An:\s", r"\n\s*Betreff:\s",
    r"\n\s*De\s*:\s", r"\n\s*Envoyé\s*:\s", r"\n\s*À\s*:\s", r"\n\s*Objet\s*:\s",
    r"\n\s*-----Original Message-----",
]
THREAD_MARKERS_RE = re.compile("|".join(THREAD_MARKERS), re.IGNORECASE)

UNDERSCORE_LINE_RE = re.compile(r"\n_{10,}\s*\n")

NOISE_LINE_RE = re.compile(
    r"(confidential|privileged|do not click|caution:|unauthorized|policy|datenschutz|privacy|disclaimer)",
    re.IGNORECASE
)

ID_KEYWORDS_RE = re.compile(
    r"(hrb|ust-?id|vat|uid|register|registernummer|tax|u\.st\.|ustnr|"
    r"rcs|siret|tva|code\s*ape|ape\s*:|intra|n°tva)",
    re.IGNORECASE
)

FAX_RE = re.compile(
    r"\b(fax|f\s*:|f\s+\+?|telefax|facsimile|\(fax\))\b",
    re.IGNORECASE
)

FAX_NUMBER_CONTEXT_RE = re.compile(
    r"\b(fax|telefax|facsimile)\b.*\d{6,}",
    re.IGNORECASE
)

PHONE_LABEL_RE = re.compile(
    r"\b(tel|phone|mobile|cell|cel|direct|d:|t:|m:)\b",
    re.IGNORECASE
)

PHONE_CANDIDATE_RE = re.compile(
    r"""
    (?:
        (?:\+?\d{1,3}[\s\-/\.]?)?
        (?:\(?\d{1,4}\)?[\s\-/\.]?)?
        \d{2,4}[\s\-/\.]?\d{2,4}(?:[\s\-/\.]?\d{2,6})?
        |
        (?:\+?\d{1,3}[\s\-/\.]?)
        \d{1,4}[\s\-/\.]?
        \d{5,9}
    )
    """,
    re.VERBOSE
)

DATE_REGEX = re.compile(
    r"""
    \b(
        \d{1,2}[\/\-\.]\d{1,2}[\/\-\.]\d{2,4} |
        \d{4}[\/\-\.]\d{1,2}[\/\-\.]\d{1,2}
    )\b
    """,
    re.VERBOSE
)

DATE_CONTEXT_RE = re.compile(
    r"\b(holiday|closed|schedule|resume|break|opening|hours|business hours)\b",
    re.IGNORECASE
)

SE_PHONE_BLACKLIST = {
    "16178655280",
    "16692443065",
}


def normalize_phone_digits(p: str) -> str:
    return re.sub(r"\D", "", p)


def digits_count(s: str) -> int:
    return len(re.sub(r"\D", "", s))


def clean_text_basic(s: str) -> str:
    if not s:
        return ""
    s = html.unescape(s)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    return s


def read_eml(path: str):
    with open(path, "rb") as f:
        msg = BytesParser(policy=policy.default).parse(f)

    from_header = msg.get("From", "") or ""

    body_part = msg.get_body(preferencelist=("plain",))
    if body_part:
        body_text = body_part.get_content()
    else:
        html_part = msg.get_body(preferencelist=("html",))
        body_text = html_part.get_content() if html_part else ""

    return clean_text_basic(body_text or ""), from_header


def read_msg(path: str):
    if extract_msg is None:
        raise RuntimeError("extract_msg is not installed.")
    m = extract_msg.Message(path)
    body = m.body or ""
    from_header = getattr(m, "sender", "") or getattr(m, "sender_email", "") or ""
    return clean_text_basic(body), from_header


def extract_latest_reply(body: str) -> str:
    if not body:
        return ""

    m = UNDERSCORE_LINE_RE.search(body)
    if m:
        body = body[:m.start()]

    m_on = ON_WROTE_RE.search(body)
    if m_on:
        body = body[:m_on.start()]

    m2 = THREAD_MARKERS_RE.search(body)
    if m2:
        body = body[:m2.start()]

    return body.strip()


def signature_window(latest: str, max_lines: int = 40) -> str:
    lines = [ln.rstrip() for ln in latest.split("\n")]
    tail = lines[-max_lines:]

    cleaned = []
    for ln in tail:
        s = ln.strip()
        if not s:
            continue
        if NOISE_LINE_RE.search(s):
            continue
        cleaned.append(s)

    return "\n".join(cleaned)


COUNTRY_CODE_HINT_RE = re.compile(
    r"(?:^|\b)(?:tel|phone|mobile|cell|cel|direct|t|m|p)\s*[:\-\s]*"
    r"(\+?\d{1,3})(?=[\s\-/\.]|\()",
    re.IGNORECASE
)

PLUS_COUNTRY_AT_LINE_START_RE = re.compile(r"^\s*(\+\d{1,3})\b")


def extract_country_code_hint(line: str) -> str:
    if not line:
        return ""

    m0 = PLUS_COUNTRY_AT_LINE_START_RE.search(line)
    if m0:
        return m0.group(1)

    m1 = COUNTRY_CODE_HINT_RE.search(line)
    if m1:
        cc = m1.group(1)
        if cc and not cc.startswith("+"):
            cc = "+" + cc
        return cc

    m2 = re.search(r"\b(?:tel|t|phone|p)\s*[:\-\s]*([0-9]{1,3})\b", line, re.IGNORECASE)
    if m2:
        return "+" + m2.group(1)

    return ""


def maybe_prepend_country_code(phone: str, country_code: str) -> str:
    if not phone:
        return phone
    if phone.strip().startswith("+"):
        return phone.strip()
    if not country_code:
        return phone.strip()

    ph_digits = normalize_phone_digits(phone)
    cc_digits = normalize_phone_digits(country_code)
    if cc_digits and ph_digits.startswith(cc_digits):
        return phone.strip()

    return f"{country_code} {phone.strip()}"


def find_phones_from_signature(sig_text: str) -> str:
    if not sig_text:
        return ""

    lines = sig_text.split("\n")
    candidates = []

    def add_from_line(line: str):
        parts = re.split(r"[|;/]", line)
        for part in parts:
            part = part.strip()
            if not part:
                continue

            part_l = part.lower()

            if FAX_RE.search(part_l):
                continue
            if FAX_NUMBER_CONTEXT_RE.search(part_l):
                continue
            if ID_KEYWORDS_RE.search(part_l):
                continue
            if DATE_CONTEXT_RE.search(part_l):
                continue

            cc_hint = extract_country_code_hint(part)

            for m in PHONE_CANDIDATE_RE.findall(part):
                ph = m.strip()

                if DATE_REGEX.search(ph):
                    continue

                dc = digits_count(ph)
                if dc < 8 or dc > 16:
                    continue

                if not (ph.startswith("+") or any(ch in ph for ch in [" ", "-", "/", "(", ")", "."])):
                    continue

                ph = maybe_prepend_country_code(ph, cc_hint)
                candidates.append(ph)

    for ln in lines:
        if PHONE_LABEL_RE.search(ln):
            add_from_line(ln)

    if not candidates:
        for ln in lines:
            add_from_line(ln)

    seen = set()
    uniq = []
    for c in candidates:
        key = re.sub(r"\s+", " ", c).strip()
        if normalize_phone_digits(key) in SE_PHONE_BLACKLIST:
            continue
        if key not in seen:
            seen.add(key)
            uniq.append(key)

    return " | ".join(uniq)


# =========================
# OCR
# =========================

def ensure_tesseract_ready():
    if not OCR_AVAILABLE:
        raise RuntimeError("OCR libraries not installed. Install: pillow, pytesseract")
    if os.path.exists(TESSERACT_CMD):
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD


def extract_images_from_eml(path: str):
    imgs = []
    with open(path, "rb") as f:
        msg = BytesParser(policy=policy.default).parse(f)

    for part in msg.walk():
        ctype = (part.get_content_type() or "").lower()
        if ctype.startswith("image/"):
            payload = part.get_payload(decode=True)
            if payload:
                imgs.append(payload)
    return imgs


def extract_images_from_msg(path: str):
    imgs = []
    if extract_msg is None:
        return imgs

    m = extract_msg.Message(path)
    try:
        attachments = m.attachments
    except Exception:
        attachments = []

    for att in attachments:
        try:
            long_name = (getattr(att, "longFilename", "") or "").lower()
            short_name = (getattr(att, "shortFilename", "") or "").lower()
            name = long_name or short_name
            if any(name.endswith(ext) for ext in [".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".gif", ".webp"]):
                data = att.data
                if data:
                    imgs.append(data)
        except Exception:
            continue

    return imgs


def ocr_images_to_text(image_bytes_list):
    if not image_bytes_list:
        return ""
    ensure_tesseract_ready()

    texts = []
    for b in image_bytes_list:
        try:
            img = Image.open(io.BytesIO(b))
            img = img.convert("L")
            img = img.resize((img.width * 2, img.height * 2))
            txt = pytesseract.image_to_string(img)
            if txt and txt.strip():
                texts.append(txt)
        except Exception:
            continue

    return "\n".join(texts)


def merge_phone_strings(a: str, b: str) -> str:
    def to_list(s):
        if not s:
            return []
        return [x.strip() for x in s.split("|") if x.strip()]

    merged = []
    seen = set()
    for item in to_list(a) + to_list(b):
        key = re.sub(r"\s+", " ", item).strip()
        if normalize_phone_digits(key) in SE_PHONE_BLACKLIST:
            continue
        if key not in seen:
            seen.add(key)
            merged.append(key)

    return " | ".join(merged)


def sender_email_from_header(from_header: str) -> str:
    _, email = parseaddr(from_header or "")
    return email.strip()


def sender_name_from_header(from_header: str, sender_email: str) -> str:
    name, _ = parseaddr(from_header or "")
    name = (name or "").strip().strip('"').strip("'")
    if name:
        return name
    if sender_email:
        base = sender_email.split("@")[0]
        base = re.sub(r"[._\-]+", " ", base).strip()
        return base.title()
    return ""


def process_email_file(path: str, use_ocr: bool = False):
    if path.lower().endswith(".eml"):
        body, from_header = read_eml(path)
    elif path.lower().endswith(".msg"):
        body, from_header = read_msg(path)
    else:
        raise ValueError("Unsupported file type")

    latest = extract_latest_reply(body)
    sig = signature_window(latest)

    sender_email = sender_email_from_header(from_header)
    sender_name = sender_name_from_header(from_header, sender_email)
    sender_phones = find_phones_from_signature(sig)

    if use_ocr:
        try:
            if path.lower().endswith(".eml"):
                imgs = extract_images_from_eml(path)
            else:
                imgs = extract_images_from_msg(path)

            ocr_text = ocr_images_to_text(imgs)
            ocr_phones = find_phones_from_signature(ocr_text) if ocr_text else ""
            sender_phones = merge_phone_strings(sender_phones, ocr_phones)
        except Exception:
            pass

    return sender_name, sender_email, sender_phones


def process_excel_input(input_excel: str, output_excel: str, use_ocr: bool = False):
    wb_in = load_workbook(input_excel)
    ws_in = wb_in.active

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Results"
    ws_out.append(["File Path", "Sender Name", "Sender Email", "Sender Phones", "Status"])

    for row in ws_in.iter_rows(min_row=2):
        raw_path = row[4].value  # Column E

        if raw_path is None or str(raw_path).strip() == "" or str(raw_path).strip().upper() == "#N/A":
            ws_out.append(["#N/A", "", "", "", "No path provided"])
            continue

        path = str(raw_path).strip()

        if not os.path.exists(path):
            ws_out.append([path, "", "", "", "File not found"])
            continue

        try:
            name, email, phones = process_email_file(path, use_ocr=use_ocr)
            ws_out.append([path, name, email, phones, "OK"])
        except Exception as e:
            ws_out.append([path, "", "", "", f"Error: {e}"])

    wb_out.save(output_excel)


# =========================
# GUI
# =========================

def browse_input_excel():
    path = filedialog.askopenfilename(
        title="Select input Excel file",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if path:
        input_var.set(path)
        default_out = os.path.join(os.path.dirname(path), "sender_results.xlsx")
        output_var.set(default_out)


def browse_output_excel():
    path = filedialog.asksaveasfilename(
        title="Save output Excel file as",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialfile="sender_results.xlsx"
    )
    if path:
        output_var.set(path)


def run_tool():
    input_excel = input_var.get().strip()
    output_excel = output_var.get().strip()
    use_ocr = bool(ocr_var.get())

    if not input_excel:
        messagebox.showerror("Error", "Please select an input Excel file.")
        return
    if not os.path.exists(input_excel):
        messagebox.showerror("Error", "Input Excel file not found.")
        return
    if not output_excel:
        messagebox.showerror("Error", "Please choose an output Excel file path.")
        return

    if use_ocr and not OCR_AVAILABLE:
        messagebox.showerror(
            "OCR not available",
            "OCR is enabled but Pillow/pytesseract is not installed.\n\nRun:\npy -m pip install pillow pytesseract"
        )
        return

    try:
        process_excel_input(input_excel, output_excel, use_ocr=use_ocr)
        messagebox.showinfo("Done", f"Extraction completed!\n\nOutput saved to:\n{output_excel}")
    except PermissionError:
        messagebox.showerror(
            "Permission Error",
            "Cannot write the output file.\n\nPlease close the Excel output file if it is open, then try again."
        )
    except Exception as e:
        messagebox.showerror("Error", f"Unexpected error:\n{e}")


class LargeCheckbox(tk.Canvas):
    def __init__(self, parent, variable, size=22):
        super().__init__(parent, width=size, height=size,
                         bg=SE_BLACK, highlightthickness=0, cursor="hand2")
        self.size = size
        self.variable = variable

        self.rect = self.create_rectangle(
            2, 2, size - 2, size - 2,
            outline=SE_MUTED, width=2, fill=SE_DARK_CARD
        )
        self.check = self.create_text(
            size // 2, size // 2, text="✓",
            fill=SE_YELLOW, font=("Arial", size - 8, "bold"),
            state="hidden"
        )

        self.bind("<Button-1>", self.toggle)
        self._update_visual()

    def toggle(self, event=None):
        self.variable.set(0 if self.variable.get() else 1)
        self._update_visual()

    def _update_visual(self):
        if self.variable.get():
            self.itemconfigure(self.rect, outline=SE_YELLOW, fill=SE_BLACK)
            self.itemconfigure(self.check, state="normal")
        else:
            self.itemconfigure(self.rect, outline=SE_MUTED, fill=SE_DARK_CARD)
            self.itemconfigure(self.check, state="hidden")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("SiliconExpert | Phone Extractor")
    root.geometry("820x340")
    root.configure(bg=SE_BLACK)
    root.resizable(False, False)

    input_var = tk.StringVar()
    output_var = tk.StringVar()
    ocr_var = tk.IntVar(value=0)

    # Header (NO LOGO)
    header = tk.Frame(root, bg=SE_BLUE, height=70)
    header.pack(fill="x")
    header.pack_propagate(False)

    header_inner = tk.Frame(header, bg=SE_BLUE)
    header_inner.pack(fill="x", padx=20, expand=True)

    dot_canvas = tk.Canvas(header_inner, width=14, height=14, bg=SE_BLUE, highlightthickness=0)
    dot_canvas.create_oval(2, 2, 12, 12, fill=SE_YELLOW, outline="")
    dot_canvas.pack(side="left", padx=(0, 6), pady=10)

    tk.Label(
        header_inner, text="SiliconExpert", fg=SE_WHITE, bg=SE_BLUE,
        font=("Segoe UI", 18, "bold")
    ).pack(side="left", pady=10)

    tk.Label(
        header_inner, text="Phone Extractor", fg=SE_YELLOW, bg=SE_BLUE,
        font=("Segoe UI", 16, "bold")
    ).pack(side="left", padx=(16, 0), pady=10)

    # Body card
    card = tk.Frame(root, bg=SE_DARK_CARD, highlightbackground=SE_BORDER, highlightthickness=1)
    card.pack(fill="both", expand=True, padx=20, pady=(16, 20))

    body = tk.Frame(card, bg=SE_DARK_CARD)
    body.pack(fill="both", expand=True, padx=24, pady=16)

    LABEL_FONT = ("Segoe UI", 11)
    ENTRY_FONT = ("Segoe UI", 10)
    BTN_FONT = ("Segoe UI", 10, "bold")

    tk.Label(body, text="Input Excel (.xlsx):", fg=SE_WHITE, bg=SE_DARK_CARD,
             font=LABEL_FONT, anchor="w", width=20).grid(row=0, column=0, sticky="w", pady=(0, 10))

    tk.Entry(body, textvariable=input_var, width=55, font=ENTRY_FONT,
             bg=SE_WHITE, fg=SE_BLACK, relief="flat", bd=0).grid(row=0, column=1, padx=(0, 10), pady=(0, 10), ipady=4)

    tk.Button(body, text="Browse", command=browse_input_excel,
              bg=SE_BLUE, fg=SE_WHITE, font=BTN_FONT, width=9, relief="flat",
              activebackground="#0960C0", activeforeground=SE_WHITE, cursor="hand2"
              ).grid(row=0, column=2, pady=(0, 10))

    tk.Label(body, text="Output Excel (.xlsx):", fg=SE_WHITE, bg=SE_DARK_CARD,
             font=LABEL_FONT, anchor="w", width=20).grid(row=1, column=0, sticky="w", pady=(0, 10))

    tk.Entry(body, textvariable=output_var, width=55, font=ENTRY_FONT,
             bg=SE_WHITE, fg=SE_BLACK, relief="flat", bd=0).grid(row=1, column=1, padx=(0, 10), pady=(0, 10), ipady=4)

    tk.Button(body, text="Browse", command=browse_output_excel,
              bg=SE_BLUE, fg=SE_WHITE, font=BTN_FONT, width=9, relief="flat",
              activebackground="#0960C0", activeforeground=SE_WHITE, cursor="hand2"
              ).grid(row=1, column=2, pady=(0, 10))

    ocr_frame = tk.Frame(body, bg=SE_DARK_CARD)
    ocr_frame.grid(row=2, column=1, sticky="w", pady=(4, 4))

    large_cb = LargeCheckbox(ocr_frame, variable=ocr_var, size=24)
    large_cb.pack(side="left")

    lbl = tk.Label(ocr_frame, text="Extract phones from images (OCR)", fg=SE_WHITE,
                   bg=SE_DARK_CARD, font=("Segoe UI", 10), cursor="hand2")
    lbl.pack(side="left", padx=(8, 0))
    lbl.bind("<Button-1>", lambda e: large_cb.toggle())

    tk.Button(
        body, text="Run Extraction", command=run_tool,
        bg=SE_YELLOW, fg=SE_BLACK, font=("Segoe UI", 12, "bold"),
        width=22, relief="flat", cursor="hand2",
        activebackground="#E5A800", activeforeground=SE_BLACK
    ).grid(row=3, column=1, pady=(14, 4))

    tk.Label(
        body,
        text="Tip: OCR requires Tesseract + pytesseract + Pillow installed.",
        fg=SE_MUTED, bg=SE_DARK_CARD, font=("Segoe UI", 8)
    ).grid(row=4, column=1, sticky="w", pady=(4, 0))

    root.mainloop()